#!/usr/bin/env python3 
"""
FUNCTION: Converts a CSV (tab delimited) file to an Excel xlsx file.

Copyright (c) 2012, Konrad Foerstner <konrad@foerstner.org>
Copyright (c) 2013, Ondrej Brablc <ondrej@brablc.com>

Permission to use, copy, modify, and/or distribute this software for
any purpose with or without fee is hereby granted, provided that the
above copyright notice and this permission notice appear in all
copies.

THE SOFTWARE IS PROVIDED 'AS IS' AND THE AUTHOR DISCLAIMS ALL
WARRANTIES WITH REGARD TO THIS SOFTWARE INCLUDING ALL IMPLIED
WARRANTIES OF MERCHANTABILITY AND FITNESS. IN NO EVENT SHALL THE
AUTHOR BE LIABLE FOR ANY SPECIAL, DIRECT, INDIRECT, OR CONSEQUENTIAL
DAMAGES OR ANY DAMAGES WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR
PROFITS, WHETHER IN AN ACTION OF CONTRACT, NEGLIGENCE OR OTHER
TORTIOUS ACTION, ARISING OUT OF OR IN CONNECTION WITH THE USE OR
PERFORMANCE OF THIS SOFTWARE.


Modifications by Andreas Brodbeck for Albus
"""

import argparse
import csv
import os, sys, inspect

# Use preferably local python libraries in search path
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

from openpyxl.workbook import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("input_file")
parser.add_argument("-d", "--delimiter", default="\t", help="select delimiter character")
args = parser.parse_args()

workbook = Workbook()
sheet = workbook.worksheets[0] # create_sheet()

for row_index, row in enumerate(
    csv.reader(open(args.input_file, 'r'), delimiter=args.delimiter), start=1):
    for col_index, col in enumerate(row, start=1):
        sheet.cell(row = row_index, column = col_index).value = col

workbook.save(args.input_file + ".xlsx")
