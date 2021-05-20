#!/usr/bin/env python2
"""
FUNCTION: Converts a CSV (tab delimited) file to an Excel xlsx file.

Copyright (c) 2012, Konrad Foerstner <konrad@foerstner.org>
Copyright (c) 2013, Ondrej Brablc <ondrej@brablc.com>

Permission to use, copy, modify, and/or distribute this software for
any purpose with or without fee is hereby granted, provided that the
above copyright notice and this permission notice appear in all
copies.

THE SOFTWARE IS PROVIDED 'AS IS' AND THE AUTHOR DISCLAIMS ALL
WARRANTIES WITH REGARD TO THIS SOFTWARE INCLUDING ALL IMPLIED
WARRANTIES OF MERCHANTABILITY AND FITNESS. IN NO EVENT SHALL THE
AUTHOR BE LIABLE FOR ANY SPECIAL, DIRECT, INDIRECT, OR CONSEQUENTIAL
DAMAGES OR ANY DAMAGES WHATSOEVER RESULTING FROM LOSS OF USE, DATA OR
PROFITS, WHETHER IN AN ACTION OF CONTRACT, NEGLIGENCE OR OTHER
TORTIOUS ACTION, ARISING OUT OF OR IN CONNECTION WITH THE USE OR
PERFORMANCE OF THIS SOFTWARE.


Modifications by Andreas Brodbeck for Albus
"""

import argparse
import ucsv as csv
import os, sys, inspect

# Use preferably local python libraries in search path
cmd_folder = os.path.realpath(os.path.abspath(os.path.split(inspect.getfile( inspect.currentframe() ))[0]))
if cmd_folder not in sys.path:
    sys.path.insert(0, cmd_folder)

from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument("input_file")
args = parser.parse_args()

def xstr(s):
    if s is None:
        return ''
    return s

# workbook = load_workbook(args.input_file, use_iterators = True)
# sheet = workbook.get_active_sheet()
# 
# with open((args.input_file + '.csv'), 'wb') as csvfile:
#   csvwriter = csv.writer(csvfile, delimiter=args.delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)
# 
# for i, row in enumerate(sheet.iter_rows()):
#   row = [ c.internal_value for c in row ]
#   csvwriter.writerow(row)

wb = load_workbook(args.input_file)
sh = wb.get_active_sheet()
with open((args.input_file + '.csv'), 'wb') as f:
    c = csv.writer(f,delimiter="\t")
    for r in sh.rows:
        c.writerow([xstr(cell.value) for cell in r][:50])
